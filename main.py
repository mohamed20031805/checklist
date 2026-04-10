import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os

# ─── Société Générale Design System ──────────────────────────────
SG_RED       = "#E60028"
SG_RED_DARK  = "#B0001E"
SG_RED_LIGHT = "#FFF0F2"
SG_BLACK     = "#1A1A1A"
SG_DARK      = "#2C2C2C"
SG_GRAY      = "#F4F4F4"
SG_GRAY_MID  = "#E8E8E8"
SG_GRAY_DARK = "#6B6B6B"
SG_WHITE     = "#FFFFFF"
SG_BORDER    = "#D0D0D0"
SG_TEXT      = "#1A1A1A"
SG_TEXT_MID  = "#5A5A5A"
SG_GREEN     = "#007A3D"

# Section accent colors (SG palette — reds/darks/neutrals)
SECTION_COLORS = {
    "Listed Securities":  "#E60028",
    "Fund of Funds":      "#B0001E",
    "FOREX":              "#2C2C2C",
    "Listed Derivatives": "#D4000F",
    "OTC":                "#5A0010",
    "Collateral":         "#8B0000",
    "Securities Lending": "#CC0022",
}

# ─── Data model ──────────────────────────────────────────────────
SECTIONS = {
    "Listed Securities": {
        "tasks": [
            {
                "name": "Sent target portfolio",
                "has_import": True,
                "subtasks": []
            },
            {
                "name": "Instruction method",
                "has_import": False,
                "subtasks": [
                    {
                        "name": "Swifts",
                        "fields": [
                            {"label": "Sent instructor BIC code", "type": "entry", "placeholder": "XXXXXXXXX"}
                        ]
                    },
                    {
                        "name": "SG Market",
                        "fields": [
                            {"label": "Commentary / Instructions", "type": "text", "placeholder": "Enter SG Market instructions..."}
                        ]
                    }
                ]
            }
        ]
    },
    "Fund of Funds": {
        "tasks": [
            {
                "name": "Type of Mutual Funds",
                "has_import": True,
                "type": "dropdown_conditional",
                "dropdown_label": "Fund type",
                "dropdown_values": ["French Fund", "Other Fund"],
                "subtasks": []
            },
            {
                "name": "Instruction method",
                "has_import": False,
                "subtasks": [
                    {
                        "name": "Swifts",
                        "fields": [
                            {"label": "BIC code", "type": "entry", "placeholder": "XXXXXXXXX"}
                        ]
                    },
                    {
                        "name": "SG Market",
                        "fields": [
                            {"label": "Commentary", "type": "text", "placeholder": "Enter commentary..."}
                        ]
                    }
                ]
            }
        ]
    },
    "FOREX": {
        "tasks": [
            {
                "name": "FX Counterparties",
                "has_import": False,
                "subtasks": [
                    {"name": "Spot", "fields": []},
                    {"name": "Forward", "fields": []},
                    {
                        "name": "Share Class Hedging",
                        "fields": [
                            {"label": "Sent instructor BIC code",        "type": "entry", "placeholder": "XXXXXXXXX"},
                            {"label": "Access to SG Markets Forex tool", "type": "combo", "values": ["Yes", "No"]},
                            {"label": "ISDA agreement to be sent",       "type": "combo", "values": ["Yes", "No"]}
                        ]
                    },
                    {"name": "SGABLULL",               "fields": []},
                    {"name": "LMA",                    "fields": []},
                    {"name": "SGCIB",                  "fields": []},
                    {"name": "External counterparties","fields": []}
                ]
            }
        ]
    },
    "Listed Derivatives": {
        "tasks": [
            {"name": "Set up with SG Prime",                         "has_import": False, "subtasks": []},
            {
                "name": "Account opening confirmation with SG Prime",
                "has_import": False,
                "subtasks": [
                    {"name": "Account opening confirmation with SG Prime", "fields": []},
                    {"name": "Margin call payment process", "fields": [
                        {"label": "Process type", "type": "combo",
                         "values": ["0 treasury", "Swifts", "SG Markets"]}
                    ]},
                    {"name": "0 treasury", "fields": []},
                    {"name": "Swifts",     "fields": []},
                    {"name": "SG Market",  "fields": [
                        {"label": "POA (contract for 0 treasury process)", "type": "combo", "values": ["Yes", "No"]}
                    ]}
                ]
            },
            {"name": "Set up with external broker / clearer",        "has_import": False, "subtasks": []},
            {"name": "Clearer Name confirmation",                     "has_import": False, "subtasks": []},
            {"name": "Account opening confirmation with the clearer", "has_import": False, "subtasks": []},
            {
                "name": "Margin call payment process",
                "has_import": False,
                "subtasks": [
                    {"name": "Swifts",    "fields": []},
                    {"name": "SG Market", "fields": [
                        {"label": "Account number",   "type": "entry", "placeholder": "XXXXXXXX"},
                        {"label": "Creation request", "type": "text",  "placeholder": "Details..."}
                    ]}
                ]
            }
        ]
    },
    "OTC": {
        "tasks": [
            {
                "name": "Type of OTC products",
                "has_import": False,
                "subtasks": [
                    {"name": "Booking (record keeping only)", "fields": []},
                    {"name": "Cash payment",                 "fields": []},
                    {"name": "Valuation",                    "fields": []}
                ]
            },
            {"name": "Instruction method for record keeping & cash payment", "has_import": False, "subtasks": []},
            {"name": "Inform client of the OTC process",                     "has_import": False, "subtasks": []},
            {"name": "Account set up in Simcorp",                            "has_import": False, "subtasks": []}
        ]
    },
    "Collateral": {
        "tasks": [
            {"name": "Collateral type (Securities, cash)", "has_import": False, "subtasks": []},
            {"name": "Collateral Agent",                  "has_import": False, "subtasks": []},
            {
                "name": "Instruction method",
                "has_import": False,
                "subtasks": [
                    {"name": "Swifts",    "fields": []},
                    {"name": "SG Markets","fields": []}
                ]
            },
            {"name": "CSA agreement",      "has_import": False, "subtasks": []},
            {"name": "Lieu du collatéral", "has_import": False, "subtasks": []}
        ]
    },
    "Securities Lending": {
        "tasks": [
            {
                "name": "Type of lending",
                "has_import": False,
                "subtasks": [
                    {"name": "Agent lending",    "fields": []},
                    {"name": "Principal lending", "fields": []}
                ]
            },
            {"name": "Borrower identification", "has_import": False, "subtasks": []},
            {"name": "Collateral type",         "has_import": False, "subtasks": []},
            {
                "name": "Instruction method",
                "has_import": False,
                "subtasks": [
                    {"name": "Swifts",    "fields": [
                        {"label": "BIC code",   "type": "entry", "placeholder": "XXXXXXXXX"}
                    ]},
                    {"name": "SG Markets","fields": [
                        {"label": "Commentary", "type": "text",  "placeholder": "Enter commentary..."}
                    ]}
                ]
            },
            {"name": "Revenue split agreement", "has_import": False, "subtasks": []},
            {"name": "SLA / Agreement",         "has_import": True,  "subtasks": []}
        ]
    }
}


class ChecklistApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SGSS – Onboarding Checklist")
        self.geometry("1300x820")
        self.minsize(1100, 700)
        self.configure(bg=SG_GRAY)

        # Client info (filled by us, not the client)
        self.client_name    = tk.StringVar()
        self.client_address = tk.StringVar()
        self.client_project = tk.StringVar()
        # Reference = Address + Project + Date (auto)
        self.client_date    = tk.StringVar(value=datetime.today().strftime("%d%m%Y"))

        self.section_vars          = {}
        self.task_vars             = {}
        self.subtask_vars          = {}
        self.field_vars            = {}
        self.import_paths          = {}
        self.import_full_paths     = {}
        self.dropdown_vars         = {}
        self._section_frames       = {}
        self._section_task_frames  = {}   # section -> frame that holds tasks, inserted RIGHT after header
        self._task_sub_frames      = {}
        self._subtask_field_frames = {}

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TScrollbar",
                        background=SG_GRAY_MID,
                        troughcolor=SG_GRAY,
                        bordercolor=SG_BORDER,
                        arrowcolor=SG_GRAY_DARK)
        style.configure("SG.TCombobox",
                        fieldbackground=SG_WHITE,
                        background=SG_WHITE,
                        foreground=SG_TEXT,
                        bordercolor=SG_BORDER,
                        arrowcolor=SG_RED)

    # ─── UI BUILD ────────────────────────────────────────────────
    def _build_ui(self):
        # ── Top bar ──
        topbar = tk.Frame(self, bg=SG_RED, height=56)
        topbar.pack(fill="x", side="top")
        topbar.pack_propagate(False)

        # SG logo-style mark
        logo_frame = tk.Frame(topbar, bg=SG_RED)
        logo_frame.pack(side="left", padx=20, pady=8)
        tk.Label(logo_frame, text="SG", font=("Arial", 18, "bold"),
                 bg=SG_RED, fg=SG_WHITE).pack(side="left")
        tk.Frame(logo_frame, bg=SG_WHITE, width=2, height=32).pack(side="left", padx=10)
        tk.Label(logo_frame, text="SGSS  ·  Onboarding Checklist",
                 font=("Arial", 13, "bold"), bg=SG_RED, fg=SG_WHITE).pack(side="left")

        tk.Label(topbar, text="Internal Use Only — Confidential",
                 font=("Arial", 9), bg=SG_RED, fg="#FFCCCC").pack(side="right", padx=20)

        # ── Main layout ──
        main = tk.Frame(self, bg=SG_GRAY)
        main.pack(fill="both", expand=True)

        # Left sidebar
        sidebar = tk.Frame(main, bg=SG_BLACK, width=210)
        sidebar.pack(fill="y", side="left")
        sidebar.pack_propagate(False)
        self._build_sidebar(sidebar)

        # Right content
        content = tk.Frame(main, bg=SG_GRAY)
        content.pack(fill="both", expand=True)

        # Client info card (top)
        self._build_client_card(content)

        # Scrollable form
        wrapper = tk.Frame(content, bg=SG_GRAY)
        wrapper.pack(fill="both", expand=True, padx=16, pady=(8, 0))

        self._canvas = tk.Canvas(wrapper, bg=SG_GRAY, highlightthickness=0)
        sb = ttk.Scrollbar(wrapper, orient="vertical", command=self._canvas.yview)
        self.form_frame = tk.Frame(self._canvas, bg=SG_GRAY)
        self.form_frame.bind("<Configure>",
            lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._canvas.create_window((0, 0), window=self.form_frame, anchor="nw")
        self._canvas.configure(yscrollcommand=sb.set)
        self._canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._canvas.bind_all("<MouseWheel>",
            lambda e: self._canvas.yview_scroll(-1*(e.delta//120), "units"))

        self._build_form()

        # Bottom action bar
        self._build_bottom_bar(content)

    def _build_sidebar(self, parent):
        # SG red accent bar at top
        tk.Frame(parent, bg=SG_RED, height=4).pack(fill="x")

        tk.Label(parent, text="SECTIONS", font=("Arial", 8, "bold"),
                 bg=SG_BLACK, fg="#888888", pady=0
                 ).pack(anchor="w", padx=16, pady=(16, 6))

        for sec in SECTIONS:
            color = SECTION_COLORS.get(sec, SG_RED)
            frm = tk.Frame(parent, bg=SG_BLACK)
            frm.pack(fill="x")
            accent = tk.Frame(frm, bg=SG_BLACK, width=3)
            accent.pack(side="left", fill="y")
            lbl = tk.Label(frm, text=sec, font=("Arial", 9),
                           bg=SG_BLACK, fg="#CCCCCC", cursor="hand2",
                           anchor="w", padx=12, pady=7)
            lbl.pack(side="left", fill="x", expand=True)
            lbl.bind("<Enter>",  lambda e, l=lbl, a=accent, c=color:
                     (l.config(fg=SG_WHITE, bg="#2C2C2C"),
                      a.config(bg=c)))
            lbl.bind("<Leave>",  lambda e, l=lbl, a=accent:
                     (l.config(fg="#CCCCCC", bg=SG_BLACK),
                      a.config(bg=SG_BLACK)))

        # Bottom version tag
        tk.Label(parent, text="v2.0  ·  SGSS Internal",
                 font=("Arial", 7), bg=SG_BLACK, fg="#444444"
                 ).pack(side="bottom", pady=8)

    def _build_client_card(self, parent):
        card = tk.Frame(parent, bg=SG_WHITE,
                        highlightbackground=SG_BORDER, highlightthickness=1)
        card.pack(fill="x", padx=16, pady=(12, 0))

        # Card header
        hdr = tk.Frame(card, bg=SG_RED)
        hdr.pack(fill="x")
        tk.Label(hdr, text="  Client Information",
                 font=("Arial", 10, "bold"), bg=SG_RED, fg=SG_WHITE,
                 pady=6, anchor="w").pack(side="left")
        tk.Label(hdr, text="Filled by SGSS  ",
                 font=("Arial", 8), bg=SG_RED, fg="#FFAAAA").pack(side="right")

        # Fields grid
        body = tk.Frame(card, bg=SG_WHITE, pady=10, padx=16)
        body.pack(fill="x")

        fields = [
            ("Client Name",  self.client_name,    40, 0, 0),
            ("Address",      self.client_address,  40, 0, 2),
            ("Project",      self.client_project,  30, 1, 0),
            ("Date (ddmmyyyy)", self.client_date,  12, 1, 2),
        ]
        for lbl_txt, var, width, row, col in fields:
            tk.Label(body, text=lbl_txt, font=("Arial", 8, "bold"),
                     bg=SG_WHITE, fg=SG_TEXT_MID
                     ).grid(row=row*2, column=col, sticky="w", padx=(0, 4), pady=(6, 0))
            e = tk.Entry(body, textvariable=var, width=width,
                         font=("Arial", 10), relief="flat",
                         bg=SG_GRAY, fg=SG_TEXT, insertbackground=SG_RED,
                         highlightbackground=SG_BORDER, highlightthickness=1,
                         highlightcolor=SG_RED)
            e.grid(row=row*2+1, column=col, sticky="ew", padx=(0, 20), pady=(2, 0))

        # Reference preview (auto-computed)
        ref_frame = tk.Frame(body, bg=SG_WHITE)
        ref_frame.grid(row=2, column=0, columnspan=4, sticky="w", pady=(10, 4))
        tk.Label(ref_frame, text="File name preview:", font=("Arial", 8, "bold"),
                 bg=SG_WHITE, fg=SG_TEXT_MID).pack(side="left")
        self._ref_preview = tk.Label(ref_frame, text="",
                                     font=("Arial", 9, "bold"),
                                     bg=SG_WHITE, fg=SG_RED)
        self._ref_preview.pack(side="left", padx=8)

        for var in (self.client_address, self.client_project, self.client_date):
            var.trace_add("write", lambda *a: self._update_ref_preview())
        self._update_ref_preview()

    def _update_ref_preview(self):
        addr = self.client_address.get().strip().replace(" ", "_")
        proj = self.client_project.get().strip().replace(" ", "_")
        date = self.client_date.get().strip()
        parts = [p for p in [addr, proj, date] if p]
        name  = "_".join(parts) if parts else "Onboarding_..."
        self._ref_preview.config(text=f"{name}.xlsx")

    def _build_bottom_bar(self, parent):
        bar = tk.Frame(parent, bg=SG_WHITE,
                       highlightbackground=SG_BORDER, highlightthickness=1)
        bar.pack(fill="x", padx=16, pady=12)

        tk.Button(bar, text="↺  Reset",
                  font=("Arial", 9), bg=SG_GRAY, fg=SG_TEXT_MID,
                  relief="flat", cursor="hand2", padx=14, pady=8,
                  activebackground=SG_GRAY_MID,
                  command=self._reset_form
                  ).pack(side="right", padx=(6, 12), pady=8)

        btn = tk.Button(bar, text="⬇  Export to Excel",
                        font=("Arial", 11, "bold"),
                        bg=SG_RED, fg=SG_WHITE,
                        relief="flat", cursor="hand2",
                        padx=24, pady=8,
                        activebackground=SG_RED_DARK,
                        command=self._export_excel)
        btn.pack(side="right", pady=8)

        tk.Label(bar, text="Export generates one Excel file with one sheet per imported document.",
                 font=("Arial", 8), bg=SG_WHITE, fg=SG_GRAY_DARK
                 ).pack(side="left", padx=12)

    def _build_form(self):
        for sec_name, sec_data in SECTIONS.items():
            self._build_section(self.form_frame, sec_name, sec_data)

    # ─── SECTION ─────────────────────────────────────────────────
    def _build_section(self, parent, sec_name, sec_data):
        color = SECTION_COLORS.get(sec_name, SG_RED)

        # Outer wrapper keeps header + tasks together
        outer = tk.Frame(parent, bg=SG_GRAY)
        outer.pack(fill="x", pady=(8, 0))

        # Section header
        hdr = tk.Frame(outer, bg=color)
        hdr.pack(fill="x")

        # Left: colored bar + name
        left = tk.Frame(hdr, bg=color)
        left.pack(side="left", fill="x", expand=True)
        tk.Frame(left, bg=SG_WHITE, width=3).pack(side="left", fill="y")
        tk.Label(left, text=f"  {sec_name.upper()}",
                 font=("Arial", 10, "bold"), bg=color, fg=SG_WHITE,
                 pady=9, anchor="w").pack(side="left")

        # Right: Expected toggle
        right = tk.Frame(hdr, bg=color, padx=12)
        right.pack(side="right")
        tk.Label(right, text="Expected:", font=("Arial", 8),
                 bg=color, fg="#FFCCCC").pack(side="left", padx=(0, 6))

        var = tk.StringVar(value="No")
        self.section_vars[sec_name] = var

        for val, active_bg in [("Yes", "#FFFFFF"), ("No", "#FFCCCC")]:
            rb = tk.Radiobutton(right, text=val, variable=var, value=val,
                                font=("Arial", 9, "bold"),
                                bg=color, fg=SG_WHITE,
                                selectcolor=color,
                                activebackground=color,
                                command=lambda s=sec_name: self._toggle_section(s))
            rb.pack(side="left", padx=2)

        # Tasks frame — inserted RIGHT AFTER the header inside the same outer wrapper
        tasks_frame = tk.Frame(outer, bg=SG_WHITE,
                               highlightbackground=SG_BORDER, highlightthickness=1)
        # Do NOT pack yet — will show/hide on toggle
        self._section_frames[sec_name]      = tasks_frame
        self._section_task_frames[sec_name] = (outer, tasks_frame)

        for task in sec_data["tasks"]:
            self._build_task(tasks_frame, sec_name, task, color)

    def _toggle_section(self, sec_name):
        _, tasks_frame = self._section_task_frames[sec_name]
        if self.section_vars[sec_name].get() == "Yes":
            tasks_frame.pack(fill="x")   # appears RIGHT after header, inside outer
        else:
            tasks_frame.pack_forget()

    # ─── TASK ────────────────────────────────────────────────────
    def _build_task(self, parent, sec_name, task, color):
        task_name   = task["name"]
        key         = (sec_name, task_name)
        is_dropdown = task.get("type") == "dropdown_conditional"

        # Task row
        row = tk.Frame(parent, bg=SG_WHITE)
        row.pack(fill="x", padx=0)

        # Left accent line (colored)
        tk.Frame(row, bg=color, width=3).pack(side="left", fill="y")

        inner = tk.Frame(row, bg=SG_WHITE, pady=7, padx=10)
        inner.pack(side="left", fill="x", expand=True)

        var = tk.BooleanVar()
        self.task_vars[key] = var

        # Checkbox styled
        cb = tk.Checkbutton(inner, variable=var,
                            bg=SG_WHITE, activebackground=SG_WHITE,
                            selectcolor=SG_WHITE,
                            fg=SG_RED, activeforeground=SG_RED,
                            command=lambda k=key: self._toggle_task(k))
        cb.pack(side="left")

        tk.Label(inner, text=task_name, font=("Arial", 10, "bold"),
                 bg=SG_WHITE, fg=SG_TEXT).pack(side="left", padx=(4, 16))

        # Dropdown (Fund of Funds type)
        if is_dropdown:
            dd_var = tk.StringVar(value="")
            self.dropdown_vars[key] = dd_var
            tk.Label(inner, text=task.get("dropdown_label", "Type") + ":",
                     font=("Arial", 8), bg=SG_WHITE,
                     fg=SG_TEXT_MID).pack(side="left")
            ttk.Combobox(inner, textvariable=dd_var,
                         values=task.get("dropdown_values", []),
                         width=14, font=("Arial", 9),
                         state="readonly", style="SG.TCombobox"
                         ).pack(side="left", padx=(4, 12))

        # Import button
        if task.get("has_import"):
            pv = tk.StringVar(value="")
            fp = tk.StringVar(value="")
            self.import_paths[key]      = pv
            self.import_full_paths[key] = fp
            btn = tk.Button(inner, text="📎 Attach document",
                            font=("Arial", 8), bg=SG_RED_LIGHT,
                            fg=SG_RED, relief="flat", cursor="hand2",
                            padx=8, pady=3,
                            activebackground="#FFE0E5",
                            command=lambda p=pv, f=fp: self._pick_file(p, f))
            btn.pack(side="left")
            tk.Label(inner, textvariable=pv, font=("Arial", 8),
                     bg=SG_WHITE, fg=SG_GREEN, wraplength=280
                     ).pack(side="left", padx=6)

        # Subtasks frame — right after this task row, inside same parent
        if task.get("subtasks"):
            sf = tk.Frame(parent, bg="#FFF8F8")
            sf.pack(fill="x")
            sf.pack_forget()
            self._task_sub_frames[key] = sf
            for sub in task["subtasks"]:
                self._build_subtask(sf, sec_name, task_name, sub, color)

        # Thin separator
        tk.Frame(parent, bg=SG_GRAY_MID, height=1).pack(fill="x")

    def _toggle_task(self, key):
        if key in self._task_sub_frames:
            if self.task_vars[key].get():
                self._task_sub_frames[key].pack(fill="x")
            else:
                self._task_sub_frames[key].pack_forget()

    # ─── SUBTASK ─────────────────────────────────────────────────
    def _build_subtask(self, parent, sec_name, task_name, sub, color):
        sub_name = sub["name"]
        key      = (sec_name, task_name, sub_name)

        row = tk.Frame(parent, bg="#FFF8F8")
        row.pack(fill="x")
        tk.Frame(row, bg=color, width=3).pack(side="left", fill="y")
        tk.Frame(row, bg=SG_RED_LIGHT, width=16).pack(side="left", fill="y")

        inner = tk.Frame(row, bg="#FFF8F8", pady=5, padx=8)
        inner.pack(side="left", fill="x", expand=True)

        var = tk.BooleanVar()
        self.subtask_vars[key] = var
        tk.Checkbutton(inner, variable=var, bg="#FFF8F8",
                       activebackground="#FFF8F8",
                       selectcolor="#FFF8F8",
                       fg=SG_RED, activeforeground=SG_RED,
                       command=lambda k=key: self._toggle_subtask(k)
                       ).pack(side="left")
        tk.Label(inner, text="▸", font=("Arial", 9), fg=color,
                 bg="#FFF8F8").pack(side="left")
        tk.Label(inner, text=f"  {sub_name}", font=("Arial", 9, "bold"),
                 bg="#FFF8F8", fg=SG_TEXT).pack(side="left", padx=4)

        if sub.get("fields"):
            ff = tk.Frame(parent, bg="#FEF0F0")
            ff.pack(fill="x")
            ff.pack_forget()
            self._subtask_field_frames[key] = ff
            for fld in sub["fields"]:
                self._build_field(ff, sec_name, task_name, sub_name, fld)

        tk.Frame(parent, bg=SG_RED_LIGHT, height=1).pack(fill="x")

    def _toggle_subtask(self, key):
        if key in self._subtask_field_frames:
            if self.subtask_vars[key].get():
                self._subtask_field_frames[key].pack(fill="x")
            else:
                self._subtask_field_frames[key].pack_forget()

    # ─── FIELD ───────────────────────────────────────────────────
    def _build_field(self, parent, sec_name, task_name, sub_name, fld):
        label = fld["label"]
        ftype = fld.get("type", "entry")
        key   = (sec_name, task_name, sub_name, label)

        row = tk.Frame(parent, bg="#FEF0F0", pady=5, padx=40)
        row.pack(fill="x")
        tk.Label(row, text=label + "  ", font=("Arial", 9),
                 bg="#FEF0F0", fg=SG_TEXT_MID, width=32,
                 anchor="w").pack(side="left")

        entry_cfg = dict(font=("Arial", 9), relief="flat",
                         bg=SG_WHITE, fg=SG_TEXT,
                         insertbackground=SG_RED,
                         highlightbackground=SG_BORDER,
                         highlightthickness=1,
                         highlightcolor=SG_RED)

        if ftype == "entry":
            var = tk.StringVar()
            self.field_vars[key] = var
            ph  = fld.get("placeholder", "")
            e   = tk.Entry(row, textvariable=var, width=26, **entry_cfg)
            e.insert(0, ph); e.config(fg="#BBBBBB")
            e.bind("<FocusIn>",  lambda ev, w=e, p=ph:
                   (w.delete(0, "end"), w.config(fg=SG_TEXT)) if w.get() == p else None)
            e.bind("<FocusOut>", lambda ev, w=e, p=ph, v=var:
                   (w.insert(0, p), w.config(fg="#BBBBBB")) if not v.get() else None)
            e.pack(side="left")

        elif ftype == "combo":
            var = tk.StringVar()
            self.field_vars[key] = var
            ttk.Combobox(row, textvariable=var,
                         values=fld.get("values", []),
                         width=14, font=("Arial", 9),
                         state="readonly", style="SG.TCombobox"
                         ).pack(side="left")

        elif ftype == "text":
            ph  = fld.get("placeholder", "")
            txt = tk.Text(row, width=32, height=3, **entry_cfg)
            txt.insert("1.0", ph); txt.config(fg="#BBBBBB")
            txt.bind("<FocusIn>",  lambda ev, w=txt, p=ph:
                     (w.delete("1.0", "end"), w.config(fg=SG_TEXT))
                     if w.get("1.0", "end-1c") == p else None)
            txt.bind("<FocusOut>", lambda ev, w=txt, p=ph:
                     (w.delete("1.0", "end"), w.insert("1.0", p), w.config(fg="#BBBBBB"))
                     if not w.get("1.0", "end-1c") else None)
            txt.pack(side="left")
            self.field_vars[key] = txt

    # ─── HELPERS ─────────────────────────────────────────────────
    def _pick_file(self, path_var, full_path_var):
        path = filedialog.askopenfilename(
            title="Select document",
            filetypes=[("All files","*.*"),("PDF","*.pdf"),
                       ("Excel","*.xlsx *.xls"),("Word","*.docx"),
                       ("Images","*.png *.jpg *.jpeg")]
        )
        if path:
            path_var.set(os.path.basename(path))
            full_path_var.set(path)

    def _reset_form(self):
        if not messagebox.askyesno("Reset", "Reset all fields?"):
            return
        for v in [self.client_name, self.client_address,
                  self.client_project]:
            v.set("")
        self.client_date.set(datetime.today().strftime("%d%m%Y"))
        for v in self.section_vars.values():  v.set("No")
        for v in self.task_vars.values():     v.set(False)
        for v in self.subtask_vars.values():  v.set(False)
        for _, tf in self._section_task_frames.values():
            tf.pack_forget()

    def _get_field_value(self, key):
        w = self.field_vars.get(key)
        if w is None: return ""
        if isinstance(w, tk.Text):
            val = w.get("1.0", "end-1c")
            for sec, sd in SECTIONS.items():
                for task in sd["tasks"]:
                    for sub in task.get("subtasks", []):
                        for fld in sub.get("fields", []):
                            if (sec, task["name"], sub["name"], fld["label"]) == key:
                                return "" if val == fld.get("placeholder","") else val
            return val
        return w.get()

    def _make_filename(self):
        addr = self.client_address.get().strip().replace(" ", "_")
        proj = self.client_project.get().strip().replace(" ", "_")
        date = self.client_date.get().strip()
        parts = [p for p in [addr, proj, date] if p]
        return ("_".join(parts) if parts else
                f"Onboarding_{datetime.today().strftime('%Y%m%d')}") + ".xlsx"

    # ─── EXCEL EXPORT ────────────────────────────────────────────
    def _export_excel(self):
        if not self.client_name.get().strip():
            messagebox.showwarning("Missing info", "Please enter the client name.")
            return

        default_name = self._make_filename()
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")],
            initialfile=default_name
        )
        if not save_path:
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Onboarding Checklist"

        # Style helpers
        def hf(sz=11, bold=True, color="FFFFFF"):
            return Font(name="Arial", size=sz, bold=bold, color=color)
        def cf(sz=10, bold=False, color="1A1A1A"):
            return Font(name="Arial", size=sz, bold=bold, color=color)
        def xf(h):
            return PatternFill("solid", fgColor=h.lstrip("#"))
        thin = Side(style="thin", color="D0D0D0")
        def brd():
            return Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        for col, w in zip("ABCDEFG", [22, 34, 12, 12, 34, 26, 18]):
            ws.column_dimensions[col].width = w

        r = 1
        # Title row
        ws.merge_cells(f"A{r}:G{r}")
        ws[f"A{r}"] = "SOCIÉTÉ GÉNÉRALE SECURITIES SERVICES  –  Onboarding Checklist"
        ws[f"A{r}"].font = hf(13); ws[f"A{r}"].fill = xf(SG_RED)
        ws[f"A{r}"].alignment = ctr; ws.row_dimensions[r].height = 34
        r += 1

        # SG sub-header
        ws.merge_cells(f"A{r}:G{r}")
        ws[f"A{r}"] = "Internal Document — Confidential"
        ws[f"A{r}"].font = Font(name="Arial", size=8, color="CC0000")
        ws[f"A{r}"].fill = xf("#FFF0F2")
        ws[f"A{r}"].alignment = ctr; ws.row_dimensions[r].height = 14
        r += 1

        # Client info block
        info_fields = [
            ("Client Name",  self.client_name.get()),
            ("Address",      self.client_address.get()),
            ("Project",      self.client_project.get()),
            ("Date",         self.client_date.get()),
            ("Reference ID", f"{self.client_address.get().strip()}_{self.client_project.get().strip()}_{self.client_date.get().strip()}"),
            ("Generated at", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]
        for lbl, val in info_fields:
            ws.merge_cells(f"A{r}:C{r}"); ws[f"A{r}"] = lbl
            ws[f"A{r}"].font = cf(9, True, "5A5A5A"); ws[f"A{r}"].fill = xf("#F4F4F4")
            ws[f"A{r}"].alignment = lft
            ws.merge_cells(f"D{r}:G{r}"); ws[f"D{r}"] = val
            ws[f"D{r}"].font = cf(9, True); ws[f"D{r}"].fill = xf("#F4F4F4")
            ws[f"D{r}"].alignment = lft; ws.row_dimensions[r].height = 16
            r += 1

        r += 1
        # Column headers
        headers = ["Section", "Task / Sub-task", "Expected", "Status",
                   "Details / BIC / Commentary", "Fund Type / Document", "Notes"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(r, col, h)
            c.font = hf(9); c.fill = xf(SG_BLACK)
            c.alignment = ctr; c.border = brd()
        ws.row_dimensions[r].height = 18; r += 1

        doc_idx = 1

        for sec_name, sec_data in SECTIONS.items():
            expected = self.section_vars[sec_name].get()
            sec_hex  = SECTION_COLORS.get(sec_name, SG_RED).lstrip("#")
            light    = "FFF0F2"

            for ti, task in enumerate(sec_data["tasks"]):
                task_name = task["name"]
                tkey      = (sec_name, task_name)
                done      = self.task_vars.get(tkey, tk.BooleanVar()).get()
                imp_name  = self.import_paths.get(tkey, tk.StringVar()).get()
                imp_full  = self.import_full_paths.get(tkey, tk.StringVar()).get()
                dd_val    = self.dropdown_vars.get(tkey, tk.StringVar()).get()
                sec_lbl   = sec_name if ti == 0 else ""

                doc_ref = ""
                if imp_full and os.path.exists(imp_full):
                    sname   = f"Doc_{doc_idx}"
                    doc_ref = f"→ Feuille '{sname}'"
                    self._add_doc_sheet(wb, sname, imp_full, sec_name, task_name, imp_name)
                    doc_idx += 1

                col_f = doc_ref if doc_ref else (dd_val if dd_val else "")

                if task.get("subtasks"):
                    for col, v in enumerate([sec_lbl, task_name, expected,
                                             "✔" if done else "", "", col_f, ""], 1):
                        c = ws.cell(r, col, v)
                        c.font = cf(9, True, sec_hex if col == 1 else "1A1A1A")
                        c.fill = xf(light); c.alignment = lft; c.border = brd()
                    ws.row_dimensions[r].height = 18; r += 1

                    for sub in task["subtasks"]:
                        skey  = (sec_name, task_name, sub["name"])
                        sdone = self.subtask_vars.get(skey, tk.BooleanVar()).get()
                        parts = []
                        for fld in sub.get("fields", []):
                            fk  = (sec_name, task_name, sub["name"], fld["label"])
                            val = self._get_field_value(fk)
                            if val: parts.append(f"{fld['label']}: {val}")
                        det = "\n".join(parts)
                        for col, v in enumerate(["", f"    ▸  {sub['name']}", expected,
                                                  "✔" if sdone else "", det, "", ""], 1):
                            c = ws.cell(r, col, v)
                            c.font = cf(9); c.fill = xf("FAFAFA")
                            c.alignment = lft; c.border = brd()
                        ws.row_dimensions[r].height = max(16, det.count("\n")*14+16)
                        r += 1
                else:
                    for col, v in enumerate([sec_lbl, task_name, expected,
                                             "✔" if done else "", "", col_f, ""], 1):
                        c = ws.cell(r, col, v)
                        c.font = cf(9, col == 1, sec_hex if col == 1 else "1A1A1A")
                        c.fill = xf("FFFFFF"); c.alignment = lft; c.border = brd()
                    ws.row_dimensions[r].height = 16; r += 1

            # Section separator
            ws.merge_cells(f"A{r}:G{r}")
            ws[f"A{r}"].fill = xf(sec_hex)
            ws.row_dimensions[r].height = 3; r += 1

        ws.freeze_panes = "A11"
        ws.auto_filter.ref = f"A10:G10"
        wb.save(save_path)

        n_docs = doc_idx - 1
        msg = f"Fichier exporté :\n{save_path}"
        if n_docs:
            msg += f"\n\n{n_docs} document(s) intégré(s) en feuille(s) séparée(s)."
        messagebox.showinfo("Export réussi ✔", msg)

    def _add_doc_sheet(self, wb, sheet_name, full_path, sec_name, task_name, basename):
        ws  = wb.create_sheet(title=sheet_name)
        col = SECTION_COLORS.get(sec_name, SG_RED)

        def xf(h):   return PatternFill("solid", fgColor=h.lstrip("#"))
        def font(**k): return Font(name="Arial", **k)

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Document  ·  {sec_name}  /  {task_name}"
        ws["A1"].font = font(size=11, bold=True, color="FFFFFF")
        ws["A1"].fill = xf(col)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for i, (lbl, val) in enumerate([
            ("File",        basename),
            ("Section",     sec_name),
            ("Task",        task_name),
            ("Imported on", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ], 2):
            ws[f"A{i}"] = lbl; ws[f"A{i}"].font = font(size=9, bold=True, color="5A5A5A")
            ws[f"B{i}"] = val; ws[f"B{i}"].font = font(size=9)
            ws.row_dimensions[i].height = 16

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 52

        ext = os.path.splitext(full_path)[1].lower()

        if ext in (".png", ".jpg", ".jpeg", ".bmp"):
            try:
                img = XLImage(full_path)
                img.anchor = "A7"
                for max_d, attr_w, attr_h in [(600, "width", "height"), (750, "height", "width")]:
                    if getattr(img, attr_w) > max_d:
                        ratio = max_d / getattr(img, attr_w)
                        img.width  = int(img.width  * ratio)
                        img.height = int(img.height * ratio)
                ws.add_image(img)
            except Exception as ex:
                ws["A7"] = f"⚠ Image non intégrée : {ex}"
                ws["A7"].font = font(size=9, italic=True, color="CC0000")

        elif ext in (".xlsx", ".xls"):
            try:
                src = openpyxl.load_workbook(full_path, data_only=True)
                sw  = src.active
                ws["A7"] = f"Contenu de : {basename}"
                ws["A7"].font = font(size=9, bold=True)
                ro = 8
                for src_row in sw.iter_rows(values_only=True):
                    for ci, val in enumerate(src_row, 1):
                        if val is not None:
                            ws.cell(ro, ci, str(val)).font = font(size=9)
                    ro += 1
                    if ro > 500:
                        ws.cell(ro, 1, "… (tronqué à 500 lignes)").font = font(
                            size=8, italic=True, color="888888")
                        break
            except Exception as ex:
                ws["A7"] = f"⚠ Impossible de lire : {ex}"
                ws["A7"].font = font(size=9, italic=True, color="CC0000")
        else:
            ws["A7"] = "📄 Chemin du fichier (ouvrir manuellement) :"
            ws["A7"].font = font(size=9, bold=True)
            ws["B7"] = full_path
            ws["B7"].font = font(size=9, color=SG_RED.lstrip("#"))
            ws.merge_cells("A8:F8")
            ws["A8"] = "ℹ  Les fichiers PDF et Word ne peuvent pas être intégrés directement dans Excel."
            ws["A8"].font = font(size=8, italic=True, color="888888")


if __name__ == "__main__":
    app = ChecklistApp()
    app.mainloop()
